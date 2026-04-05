from __future__ import annotations
"""
utils/excel_exporter.py
-----------------------
Export the analysed thread DataFrame to a professionally
formatted Excel workbook with two sheets:
  1. Thread Dashboard  — full table with colour-coded sentiment
  2. Summary           — KPI stats + bar chart
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from pathlib import Path


# ── Colour palette ────────────────────────────────────────────────────────────
NAVY       = "1F4E79"
WHITE      = "FFFFFF"
LIGHT_BLUE = "EBF5FB"
ALT_ROW    = "F4F9FD"
URGENT_BG  = "FDECEA"
URGENT_FX  = "C0392B"
NEG_BG     = "FEF0F0"
POS_BG     = "EAFAF1"
NEUTRAL_BG = "FAFAFA"
TOTAL_BG   = "D4E6F1"
HEADER_FX  = "2C3E50"

SENT_FILLS = {
    "urgent"  : (URGENT_BG, URGENT_FX),
    "negative": (NEG_BG,    "922B21"),
    "positive": (POS_BG,    "1E8449"),
    "neutral" : (NEUTRAL_BG, HEADER_FX),
}

DISPLAY_COLS = [
    "Email Thread",
    "Key Topic",
    "Action Items",
    "Owner",
    "Follow-ups",
    "Sentiment",
    "Email Count",
    "Latest Date",
]

COL_WIDTHS = {
    "Email Thread" : 30,
    "Key Topic"    : 28,
    "Action Items" : 38,
    "Owner"        : 20,
    "Follow-ups"   : 30,
    "Sentiment"    : 12,
    "Email Count"  : 10,
    "Latest Date"  : 14,
}


def _thin_border() -> Border:
    s = Side(border_style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)


def export_to_excel(df: pd.DataFrame, output_path: str = "email_dashboard.xlsx") -> None:
    """
    Write the thread analysis DataFrame to a formatted Excel dashboard.

    Args:
        df          : DataFrame from analyse_all_threads()
        output_path : Destination file path
    """
    df_out = df[[c for c in DISPLAY_COLS if c in df.columns]].copy()

    wb = Workbook()

    _build_dashboard_sheet(wb.active, df_out)
    _build_summary_sheet(wb.create_sheet("Summary"), df)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✓ Excel dashboard saved → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — THREAD DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def _build_dashboard_sheet(ws, df: pd.DataFrame) -> None:
    ws.title = "Thread Dashboard"
    ws.sheet_view.showGridLines = False

    # Title row
    n_cols = len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value     = "📧  Group Email Intelligence Dashboard — Enron Dataset"
    c.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle row
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c2 = ws["A2"]
    c2.value     = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  " \
                   f"Threads: {len(df)}  |  Powered by: sumy · KeyBERT · VADER · spaCy"
    c2.font      = Font(name="Arial", italic=True, size=9, color="7F8C8D")
    c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 15

    # Header row (row 3)
    hdr_font  = Font(name="Arial", bold=True, size=10, color=WHITE)
    hdr_fill  = PatternFill("solid", start_color="2980B9")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_i, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_i, value=col_name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _thin_border()
    ws.row_dimensions[3].height = 22

    # Data rows (start at row 4)
    for row_i, (_, row) in enumerate(df.iterrows(), start=4):
        sentiment = str(row.get("Sentiment", "neutral")).lower()
        bg, fx    = SENT_FILLS.get(sentiment, (NEUTRAL_BG, HEADER_FX))
        fill      = PatternFill("solid", start_color=bg)
        is_alt    = row_i % 2 == 0

        for col_i, col_name in enumerate(df.columns, start=1):
            val  = row[col_name]
            cell = ws.cell(row=row_i, column=col_i, value=str(val) if val is not None else "")
            cell.border    = _thin_border()
            cell.font      = Font(name="Arial", size=9, color=fx)

            if sentiment in ("urgent", "negative", "positive"):
                cell.fill = fill
            else:
                cell.fill = PatternFill("solid", start_color=ALT_ROW if is_alt else WHITE)

            if col_name in ("Email Count",):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "Sentiment":
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.row_dimensions[row_i].height = 52   # Tall rows for wrapped text

    # Totals row
    total_row = 4 + len(df)
    total_fill = PatternFill("solid", start_color=TOTAL_BG)
    total_font = Font(name="Arial", bold=True, size=10, color="1A5276")
    for col_i, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=total_row, column=col_i)
        cell.fill   = total_fill
        cell.font   = total_font
        cell.border = _thin_border()
        if col_name == "Email Thread":
            cell.value = f"TOTAL — {len(df)} threads analysed"
        elif col_name == "Email Count":
            col_letter = get_column_letter(col_i)
            cell.value = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.value = ""

    # Column widths
    for col_i, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = COL_WIDTHS.get(col_name, 18)

    # Freeze title + header
    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(ws, df: pd.DataFrame) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "📊  Email Group Summary Statistics"
    c.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # KPI stats
    sent_counts  = df["Sentiment"].value_counts().to_dict() if "Sentiment" in df.columns else {}
    total        = len(df)
    urgent_cnt   = int(sent_counts.get("urgent", 0))
    negative_cnt = int(sent_counts.get("negative", 0))
    positive_cnt = int(sent_counts.get("positive", 0))
    avg_emails   = round(df["Email Count"].mean(), 1) if "Email Count" in df.columns else 0
    total_emails = int(df["Email Count"].sum()) if "Email Count" in df.columns else 0

    kpis = [
        ("Total Threads",        total,        NAVY),
        ("Total Emails",         total_emails, "2471B5"),
        ("Avg Emails/Thread",    avg_emails,   "117864"),
        ("Urgent Threads",       urgent_cnt,   "C0392B"),
        ("Negative Threads",     negative_cnt, "922B21"),
        ("Positive Threads",     positive_cnt, "1E8449"),
    ]

    # Write KPI cards in a row
    for col_i, (label, value, color) in enumerate(kpis, start=1):
        # Value cell
        vc = ws.cell(row=3, column=col_i, value=value)
        vc.font      = Font(name="Arial", bold=True, size=22, color=color)
        vc.alignment = Alignment(horizontal="center")
        vc.fill      = PatternFill("solid", start_color="EBF5FB")
        vc.border    = _thin_border()
        ws.row_dimensions[3].height = 36

        # Label cell
        lc = ws.cell(row=4, column=col_i, value=label)
        lc.font      = Font(name="Arial", size=9, color="7F8C8D")
        lc.alignment = Alignment(horizontal="center")
        lc.fill      = PatternFill("solid", start_color="F8FBFD")
        lc.border    = _thin_border()
        ws.row_dimensions[4].height = 16

        ws.column_dimensions[get_column_letter(col_i)].width = 16

    # Sentiment breakdown table (for chart)
    ws["A6"] = "Sentiment Breakdown"
    ws["A6"].font = Font(name="Arial", bold=True, size=11, color="1F4E79")

    ws.cell(row=7, column=1, value="Sentiment").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=7, column=2, value="Count").font     = Font(bold=True, name="Arial", size=10)

    sent_order = ["urgent", "negative", "neutral", "positive"]
    for row_i, sent in enumerate(sent_order, start=8):
        cnt = int(sent_counts.get(sent, 0))
        ws.cell(row=row_i, column=1, value=sent.capitalize()).font = Font(name="Arial", size=10)
        ws.cell(row=row_i, column=2, value=cnt).font               = Font(name="Arial", size=10)

    # Bar chart
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Threads by Sentiment"
    chart.style = 10
    chart.y_axis.title = "Count"
    chart.width  = 15
    chart.height = 10

    data = Reference(ws, min_col=2, min_row=7, max_row=11)
    cats = Reference(ws, min_col=1, min_row=8, max_row=11)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D3")
